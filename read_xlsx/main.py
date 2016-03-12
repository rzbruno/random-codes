from openpyxl import load_workbook

FILE_NAME = 'data.xlsx'

wb = load_workbook(filename=FILE_NAME, read_only=True)
sheet_name = wb.get_sheet_names()[0]
ws = wb[sheet_name]

# wb['D18'].value
# ws.cell('A4')
# ws.cell(row = 4, column = 2)

for row in ws.rows:
    for cell in row:
        print(cell.value)
